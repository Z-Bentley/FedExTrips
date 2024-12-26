# Library Imports
import openpyxl
import os
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

# File Imports
import WeightCalculations
import CalcSortTime
import copy_excel

##### Main file for running the Trips Program #####

# Global variable to hold the Excel application instance
excel_app = None

# Create Window
def submit_data():
    data = {
        "localSchTimes": [schArrival.get(), schSortStart.get(), schSortEnd.get()],
        "localActTimes": [actArrival.get(), actSortStart.get(), actSortEnd.get()],
        "outSchTimes": [var.get() for var in schTimesVars],
        "outActTimes": [var.get() for var in truckTimesVars]
    }
    return data

# Browse for Excel File
def browseFiles():
    global filePath, templatePath, sortTimeSheet, wb

    # Chosen Excel file path
    filePath = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])

    # Predefined Excel template
    templatePath = os.path.abspath('Excel-Documents\\Sort_Time.xlsx')

    if not os.path.exists(templatePath):
        print(f"Template file not found at {templatePath}. Please check the path.")
        fileSelectorButton.config(text='Template file not found.')
        return

    if filePath:
        fileSelectorButton.config(text=f'Selected File: {filePath}')
        wb = openpyxl.load_workbook(templatePath)

        # Ensure 'sort_times' sheet exists and set it as `sortTimeSheet`
        sheetName = 'sort_times'
        if sheetName in wb.sheetnames:
            wb.remove(wb[sheetName])  # Remove if it exists
            print(f"Existing sheet '{sheetName}' has been deleted.")

        sortTimeSheet = wb.create_sheet(sheetName)  # Create fresh 'sort_times' sheet
        print(f"New sheet '{sheetName}' has been created.")
    else:
        fileSelectorButton.config(text='No file selected')
        sortTimeSheet = None

# Process data and perform calculations
def subButton():
    global sortTimeSheet, templatePath, excel_app

    if sortTimeSheet is None:
        print("No file selected or sheet not created. Please select a file first.")
        return

    # Submit Data
    data = submit_data()
    try:
        # Perform operations on `sortTimeSheet`
        CalcSortTime.calcSortTimes(sortTimeSheet, data['localSchTimes'], data['localActTimes'])
        CalcSortTime.setRootCauseDelay(sortTimeSheet, ['10856', '924 116 of this was NCING'])
        CalcSortTime.outboundTruckRoutes(sortTimeSheet, data['outSchTimes'], data['outActTimes'])

        # Save the template file
        wb.save(templatePath)

        # Copy data to clipboard using Excel
        copy_excel.copyExcel(templatePath, excel_app)
    except Exception as e:
        print(f"An error occurred in submit_data: {e}")
    finally:
        wb.save(templatePath)
        print('Workbook has been saved!')

# Close Excel when the program exits
def on_close():
    global excel_app
    if excel_app:
        copy_excel.close_excel(excel_app)
    root.destroy()

# Setup window
root = tk.Tk()
root.title("FedEx Trips Reformer")
root.geometry("600x600")

fileSelectorButton = ttk.Label(root, text='No file selected', foreground='blue')
fileSelectorButton.grid(row=1, column=0, columnspan=4, pady=5)
tk.Button(root, text="Browse", command=browseFiles).grid(row=0, column=2)

# Pre-filled data and fillable data for Aircraft arrival and Sort
schArrival = tk.StringVar(value="06:02")
schSortStart = tk.StringVar(value="06:26")
schSortEnd = tk.StringVar(value="06:46")
actArrival = tk.StringVar(value="06:46")  
actSortStart = tk.StringVar(value="06:46")  
actSortEnd = tk.StringVar(value="06:46")  

# Input fields for Aircraft and Sort times
ttk.Label(root, text="Local Sort Plan").grid(row=2, column=0, pady=(20,5))

labels_texts = ["Aircraft Arrival", "Sort Start", "Sort End"]
variables = [(schArrival, actArrival), (schSortStart, actSortStart), (schSortEnd, actSortEnd)]

for i, text in enumerate(labels_texts):
    ttk.Label(root, text=text).grid(row=i+3, column=0)
    ttk.Entry(root, textvariable=variables[i][0]).grid(row=i+3, column=1)
    ttk.Entry(root, textvariable=variables[i][1]).grid(row=i+3, column=2)

# Truck routes data
truckRoutes = ['OXD02', 'CVG10', 'CVG03', 'FFT02', 'CVG06', 'OXD04', 'LUK01', 'CVG02', 'Docs LUK77/CVG77/OXD77FFT77', 'CVG78 (DNCA)', 'FFT41 (PDJA)']
schTimes = ['06:35', '07:25', '06:45', '07:15', '06:55', '07:00', '07:10', '07:05', '06:30', '07:00', '07:20']
truckTimes = ['09:05', '09:25', '09:05', '07:22', '07:00', '07:22', '07:05', '07:25', '07:05', '07:20', '07:20']

# Variables to store scheduled and actual times
schTimesVars = [tk.StringVar(value=time) for time in schTimes]
truckTimesVars = [tk.StringVar(value=time) for time in truckTimes]

# Input fields for Truck routes
ttk.Label(root, text="Outbound Truck Routes").grid(row=5, column=0)

for i, route in enumerate(truckRoutes):
    rowNum = i + 6
    ttk.Label(root, text=route).grid(row=rowNum, column=0)
    ttk.Entry(root, textvariable=schTimesVars[i]).grid(row=rowNum, column=1)  # Scheduled times
    ttk.Entry(root, textvariable=truckTimesVars[i]).grid(row=rowNum, column=2)  # Actual times

# Submit button
ttk.Button(root, text="Submit", command=subButton).grid(row=rowNum + 1, column=1, pady=20)

# Initialize Excel at startup
excel_app = copy_excel.initialize_excel()

# Handle window close event
root.protocol("WM_DELETE_WINDOW", on_close)

root.mainloop()
