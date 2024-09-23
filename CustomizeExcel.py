from openpyxl.styles import Border, Side, PatternFill

# Border Control
def addBorder(sheet, range):
    mediumSide = Side(style='medium', color='000000')
    border = Border(
        left=mediumSide,
        right=mediumSide,
        top=mediumSide,
        bottom=mediumSide
    )

    for row in sheet[range]:
        for cell in row:
            cell.border = border

# Portion Control
def adjustCol(sheet, col):
    maxLen = 0

    # find longest cell to base expansion
    for cell in sheet[col]:
        if cell.value:
            maxLen = max(maxLen, len(str(cell.value)))
    
    adjustedWid = maxLen + 1
    sheet.column_dimensions[col].width = adjustedWid

# Color Control
def changeRowColor(sheet, rowNum, desiredColor):
    # Yellow fill
    yellowFill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Green fill
    greenFill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for cell in sheet[rowNum]:
        if desiredColor == 'green':
            cell.fill = greenFill
        elif desiredColor == 'yellow':
            cell.fill = yellowFill

