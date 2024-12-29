###### Outline of Trips program
import openpyxl

# Dictionary for ULD weights
ULD_WEIGHTS = {
    'AAD': 573,
    'AMJ': 716,
    'TRK': 750,
    'AYY': 272,
    'PMC': 260,
    'AKE': 178,
    'AAX': 865
}

# Gets ULD number from the Excel sheet
def getCanNum(cans):
    return [can.value for can in cans]


# Extracts the ULD type (first three characters)
def checkUldType(arrayOfCans):
    return [can[:3] for can in arrayOfCans if can]


# Gets the weight of a ULD type
def weightOfCan(type):
    return ULD_WEIGHTS.get(type, 0)  # Defaults to 0 if type is unknown


# Sums weights from a list of cells
def getWeight(weightOfCans):
    return sum(int(cell.value) for cell in weightOfCans if cell.value)


# Finds rows corresponding to the specified destination in column E
def getDestCans(dest, sheet):
    destCoords = []
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value == dest:
                destCoords.append(cell.row)
    return destCoords


# Gets all unique destination options from column E
def getDestOptions(sheet):
    destOptions = set()
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=5, max_col=5):  # Column E is the 5th column
        for cell in row:
            if cell.value:
                destOptions.add(cell.value)
    return list(destOptions)


# Main function
def calcWeight(sheet, dest=None):
    total_weight = 0

    try:
        # Get the total number of rows in the sheet
        num_rows = sheet.UsedRange.Rows.Count

        # Iterate through the rows
        for row in range(5, num_rows + 1):  # Assuming data starts at row 5
            dest_cell = sheet.Cells(row, 5).Value  # Column E
            weight_cell = sheet.Cells(row, 4).Value  # Column D

            # Check if the destination matches (if provided) and add weight
            if weight_cell:
                if dest is None or dest_cell == dest:
                    total_weight += int(weight_cell)

    except Exception as e:
        print(f"Error in calcWeight: {e}")

    # Format the weight with commas
    return f"{total_weight:,}"



# for a specific uld destination
# wb = openpyxl.load_workbook('Excel-Documents\\WBManifestTable_1706103354202.xlsx')
# sheet = wb['FedEx Air Ops Workbench Report']
# upperDest = 'CVGRT'
# print(calcWeight(sheet))
# print(calcWeight(sheet, upperDest))
